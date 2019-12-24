# from tkinter import *
#
# class App:
#     def __init__(self, root):
#         self.root = root
#         self.btn = Button(self.root, text="Done", command=self.command)
#         self.btn.pack()
#     def command(self):
#         self.root.destroy()
#         print("Output")
#
# root = Tk()
# app = App(root)
# root.mainloop()

import openpyxl

wb = openpyxl.load_workbook('test1.xlsx')

ws1 = wb['PersonB']

userinput1 = input("Enter the Data: ").upper()

for row in ws1.iter_rows(min_row=3, max_row=7, max_col=1):
    for cell in row:
        data = cell.value

        if data == None:
            cell.value = userinput1
            break


        else:
            print('Data here')


#Trying to Add data at the empty column and Row

# print("Number of Columns = ", ws1.max_column)
# print("Number of Rows = ", ws1.max_row)
#
#
# for i in range (3, ws1.max_row):
#     ws1.cell(row=i, column=1).value = 2
#     print(ws1.cell(row=i, column=1).value, end="  ")

wb.save('test1.xlsx')