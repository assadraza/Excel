import openpyxl
from openpyxl import *
import tkinter
from tkinter import *

# Creating Windows
window = Tk()
window.title("User Input")
window.geometry("1080x720")
window.configure(background="silver")
#####################################################################################################
# Creating Variable to Store User Input in Box1
item1 = tkinter.StringVar()

# Creating User Input Box
box1 = Entry(window, width=35, textvariable=item1)  # Creating Data Entry Box
box1.place(x=150, y=301)

entry1 = Label(window, width=20, text="*PLEASE ENTER ITEM NAME: ", foreground='black', bg='silver')  # Creating Lable
entry1.place(x=1, y=300)

def action():
    wb = openpyxl.load_workbook('test1.xlsx')
    ws1 = wb['PersonA']
    # ws1['A4'] = (box1.get())
    ws1.cell(column=1, row=ws1.max_row + 1, value=box1.get())
    wb.save('test1.xlsx')


# Creating Button

btn1 = Button(window, text="Submit", command=action)
btn1.place(x=380, y=301)

#####################################################################################################
#####################################################################################################
# Creating Variable to Store User Input in Box2
item2 = tkinter.StringVar()

# Creating User Input Box
box2 = Entry(window, width=35, textvariable=item2)  # Creating Data Entry Box
box2.place(x=150, y=350)

entry2 = Label(window, width=20, text="*PLEASE ENTER QUANTITY: ", foreground='black', bg='silver')  # Creating Lable
entry2.place(x=1, y=350)

def action():
    wb = openpyxl.load_workbook('test1.xlsx')
    ws1 = wb['PersonA']
    ws1['B4'] = (box2.get())
    wb.save('test1.xlsx')

# Creating Button

btn2 = Button(window, text="Submit", command=action)
btn2.place(x=380, y=350)

#####################################################################################################
window.mainloop()


# wb = openpyxl.load_workbook('test1.xlsx')
#
# ws1 = wb['Sheet1']

# userinput1 = input("Enter the Data: ").upper()
#
# for row in ws1.iter_rows(min_col=5):
#     for cell in row:
#         data = cell.value
#
#         if data == userinput1:
#             print('matched')
#
#         else:
#             ws1.cell(column=5, row=ws1.max_row, value=userinput1)




# userinput1 = input("Enter the Data: ")
#
# if ws1.cell(row=ws1.max_row + i, column=1).value == userinput1:
#     print("match")
# else:
#     ws1.cell(column=1, row=ws1.max_row + 1, value=userinput1)
#     print("No Match so Adding Data")

# wb.save('test1.xlsx')

# def action():
#     #entry1.configure(text=box1.get())
#     with open('test.txt', 'a+') as file:
#         file.write(box1.get())  # this will add the text at the end of the file / create if dont exist
#
#     print(file.closed)


# # Creating function
#
# def greet ():
#     print('hello')
#     print('good morning')

# To find empty or used cell in worksheet

# if ws1.cell(row = 1, column=2).value == None:
#     print("Blank")
# else:
#     print("No blank")
