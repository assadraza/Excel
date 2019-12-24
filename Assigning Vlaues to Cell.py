import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill
import xlsxwriter
from openpyxl.styles.colors import RED
from openpyxl.styles import  Fill
from openpyxl.utils import column_index_from_string

wb = openpyxl.load_workbook('data1.xlsx')
print(wb.sheetnames)

ws1 = wb.active
ws1.title = "Layer 2"
ws1.sheet_properties.tabColor = "1072BA"

#Heading and Styling

ws1['A1'].value = 'VLAN ID'
ws1['A1'].font = Font(color = "FF0005", bold=4, size=13)

ws1['B1'].value = 'ZONE'
ws1['B1'].font = Font(color = "FF0005", bold=4, size=13)

ws1['C1'].value = 'COMMENT'
ws1['C1'].font = Font(color = "FF0005", bold=4, size=13)

#Trying to Add data at the empty column and Row

print("Number of Columns = ", ws1.max_column)
print("Number of Rows = ", ws1.max_row)


for i in range (2, ws1.max_row):
    for j in range (2,ws1.max_column):
        ws1.cell(row=i, column=j).value = 1
        print(ws1.cell(row=i, column=j).value, end="  ")

    #print() #used to add new line to the columns



wb.save('data1.xlsx')