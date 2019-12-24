import openpyxl
import xlsxwriter
from openpyxl.utils import column_index_from_string

wb = openpyxl.load_workbook('Section1\data.xlsx')

# wb.sheetnames    #Get existing sheet names

# print(wb.sheetnames)
#Create Sheet

# wb.create_sheet(title='mysheet')
# wb.sheetnames
# print(wb.sheetnames)

#Create sheet at particular index

# wb.create_sheet(index=1, title="index")
#
# wb.sheetnames
# print(wb.sheetnames)

#Deleting Sheet

# std=wb.get_sheet_by_name('mysheet')   #create reference to be used to delete
# wb.remove_sheet(std)        #remove sheet using reference
#
# wb.sheetnames
# print(wb.sheetnames)

# Print Values in a Sheet at specific column
#
# sheet1 = wb['New Title']
# print(sheet1['A1'].value)
# print(sheet1['B1'].value)

# Printing Number of Max Col and Row

# sheet1 = wb['New Title']
# print(sheet1.title)
# print("Number of Columns = ", sheet1.max_column)
# print("Number of Rows = ", sheet1.max_row)

# Printing all column values in a row

# sheet1 = wb['New Title']
# for item in range (1,sheet1.max_column+1):
#     print (sheet1.cell(row=1, column=item).value, end=" ")

# Printing all row values in a column

# sheet1 = wb['New Title']
# for item in range (1, sheet1.max_row+1):
#     print(sheet1.cell(row=item, column=1).value, end=" \n")

# Printing whole sheet's column and rows

sheet1 = wb['New Title']
for i in range (1, sheet1.max_row+1):
    for j in range (1,sheet1.max_column+1):
      print(sheet1.cell(row=i, column=j).value, end="  ")
    print() #used to add new line to the columns


#wb.save('section1\data.xlsx')