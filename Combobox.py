import tkinter
from tkinter import *
from tkinter.ttk import Combobox
from tkinter.font import Font
import openpyxl

window = Tk()
window.title("User Input")
window.geometry("1080x720")

font1 = Font(family = "Times New Roman", size=16, weight="bold")
v = ["1", '2', '3', '4', 'This is Amazing']

label = Label (window, text = "Tech Gram Academy", font=font1).pack()

combo = Combobox (window, values=v)
combo.set('')

def printme ():
    value = combo.get()
    print(value)

    wb = openpyxl.load_workbook('test1.xlsx')
    ws1 = wb['PersonB']
    ws1.cell(column=1, row=ws1.max_row + 1, value=combo.get())
    wb.save('test1.xlsx')



combo.pack()
button = Button(window, text = 'Print Selection', command=printme)
button.pack()


window.mainloop()